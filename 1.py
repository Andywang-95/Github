from openpyxl import load_workbook, Workbook
import os
import cn2an
# 一
# 获取文件夹中的文件名称
path = '高效办公实战训练营\\1\第1关\static\销售数据\\'
files = os.listdir(path)
# 设定存储总表数据的空列表
total_row = []
# 设定表头数据的空列表
head_row = []
# 设定平均销量和排名的空字典
avg_dict = {}
avg_rank = {}
# 遍历销售数据资料夹的所有文件名
for file in files:
    file_path = path + file
    # 以文件名打开工作簿、工作表
    wb = load_workbook(file_path)
    ws = wb.active
    group_sales = 0
    # 获取文件名称(销售组别)以便添加进内容中
    group = file.split('.')[0]
    # 读取工作表第二行之后的内容
    for row in ws.iter_rows(min_row=2,values_only=True):
        # 将元组转换为列表以便添加、修改元素
        row = list(row)
        # 计算总销售数量
        total_sales = sum(row[2:])
        # 添加组别至列表第二个位置
        row.insert(1,group)
        # 添加总销售数量至列表末尾
        row.append(total_sales)
        # 将列表嵌套至总列表中做循环
        total_row.append(row)
        group_sales += total_sales
    # 计算小组平均销量写入字典
    if avg_dict.get(group) is None:
        avg_dict[group] = round((group_sales/(ws.max_row-1)),2)
    # 以降序排列平均销量
    avg_dict = dict(sorted(avg_dict.items(),key=lambda x:x[1],reverse=True))
    # 索引平均销量字典，添加排名到销量排名字典中
    for index, rank in enumerate(avg_dict,1):
        avg_rank[rank] = index

# 以降序的形式对整个列表重新排序，取值为每个列表的最后一个元素        
total_row = sorted(total_row,key=lambda a:a[-1],reverse=True)
# 对列表进行索引并添加到最后一位
for index,row in enumerate(total_row,start=1):
    row.append(index)
# 遍历任意销售表的第一行存成表头
for item in ws[1]:
    head_row.append(item.value)
# 添加表头数据
head_row.insert(1,'销售小组')
head_row.append('总计/瓶')
head_row.append('销售排名')

# 新建《销售总表》的工作表
wb_total = Workbook()
ws_new = wb_total.active
# 写入表头
ws_new.append(head_row)
# 写入全部值
for row in total_row:
    ws_new.append(row)
# 储存《销售总表》到对应路径
wb_total.save('高效办公实战训练营\\销售总表.xlsx')

# 二
# 设定《等级销售表》文件夹路径
path2 = '高效办公实战训练营\等级销售表\\'
# 判断路径是否存在，不存在的话新建文件夹
if not os.path.exists(path2):
    os.mkdir(path2)
# 设定每个等级的人数以及总共有几个等级
cut = 120
level = 4
# 从等级一开始新建工作表，循环
for num in range(level):
    wb_level = Workbook()
    ws_level = wb_level.active
    # 写入表头
    ws_level.append(head_row)
    # 反复切片120个值存进新的工作表中
    for row in total_row[num*cut:(num+1)*cut]:
        ws_level.append(row)
    # 储存工作簿，对文件名称进行数字-中文数字的转换
    wb_level.save(path2 + '等级{}销售表.xlsx'.format(cn2an.an2cn(num+1)))


# 三
# 打开《等级一销售表》取值
wb1 = load_workbook('高效办公实战训练营\等级销售表\等级一销售表.xlsx')
ws1 = wb1.active
# 新建《小组销售排名》工作表
wb_rank = Workbook()
ws_rank = wb_rank.active
# 设定小组优秀频次字典
group_dict = {}
# 遍历《等级一销售表》
for row in ws1.iter_rows(min_row=2,values_only=True):
    # 获取组别
    group = row[1]
    # 判断组别是否在字典中
    if group_dict.get(group) is None:
        group_dict[group] = 0
    # 该组别优秀频次加1
    group_dict[group] += 1
# 以降序排列优秀频次字典
group_dict = dict(sorted(group_dict.items(),key = lambda x:x[1],reverse=True))
# 添加表头至《小组销售排名》
ws_rank.append(['销售小组','优秀频次','频次排名','平均销量','销量排名'])
# 索引优秀频次字典（从1开始）
for index,row in enumerate(group_dict,1):    
    # 设定空列表
    l1 = []
    # 添加组别
    l1.append(row)
    # 添加优秀频次
    l1.append(group_dict[row])
    # 添加优秀频次排名（索引值）
    l1.append(index)
    # 添加该组别平均销量
    l1.append(avg_dict[row])
    # 添加该组别销量排名
    l1.append(avg_rank[row])
    # 写入工作表中
    ws_rank.append(l1)
wb_rank.save('高效办公实战训练营\小组销售排名.xlsx')