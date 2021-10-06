from openpyxl import load_workbook
import os
# 打开三个工作表
wb_schedule = load_workbook('实战演练\产量达标率\工作\生产计划表.xlsx')
ws_sc = wb_schedule.active

wb_daily = load_workbook('实战演练\产量达标率\工作\工人产量日报表.xlsx')
ws_dl = wb_daily.active

wb_record = load_workbook('实战演练\产量达标率\工作\检验记录表模板.xlsx')
ws_re = wb_record.active

# 遍历生产计划表需要复制的内容
for row_sc in ws_sc.iter_rows(min_row=3,max_col=4,values_only=True):
    # 设置计数器，用以加总各项目实际产量
    count = 0
    # 遍历日报表
    for row_dl in ws_dl.iter_rows(min_row=3,max_col=5,values_only=True):
        # 因为有相同的商品在不同车间生产，所以要判断车间和商品同时符合条件才可以进行运算
        if row_sc[1] in row_dl[0] and row_sc[2] == row_dl[1]:
            # 计算该项目每日实际产量
            count += row_dl[4]
    # 计算该项目达标率
    rate = str(round(count/row_sc[3]*100,2))+'%'
    # 将元组相加形成所需要的新元组
    list_row = row_sc + (count,rate)
    # 将列表写入模版中
    ws_re.append(list_row)
# 另存写入内容文件为新文件
wb_record.save('实战演练\产量达标率\工作\8月25日检验记录表.xlsx')

