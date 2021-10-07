# 案例 4：统计广告投放数据
# 目标：把【销售数据表格.xlsx】工作簿和【账户报表.xlsx】工作簿中对应的数据
     # 复制或经过计算记录到【广告投放数据.xlsx】工作簿中
     # 最后另存为【2020年9月11日广告投放数据.xlsx】工作簿
# 请随时查看知识库和案例练习助手，与自己编写代码的步骤和内容比对参考，训练思维
from openpyxl import load_workbook

wb_ad = load_workbook('实战演练\广告\工作\广告投放数据.xlsx')
ws_ad = wb_ad.active
wb_sale = load_workbook('实战演练\广告\工作\销售数据表格.xlsx')
ws_sl = wb_sale.active
wb_report = load_workbook('实战演练\广告\工作\账户报表.xlsx')
ws_re = wb_report.active

dict_ad = {}

for row in ws_ad.iter_rows(min_row=3,max_row=3,min_col=3,values_only=True):
    
    for r in ws_re.iter_rows(min_row=2,values_only=True):
        dict_ad['花费'] = r[6]
        dict_ad['点击量'] = r[4]
        dict_ad['点击率'] = r[5]
        dict_ad['PPC'] = round(dict_ad['花费']/dict_ad['点击量'],2)
        dict_ad['加购数'] = r[11]
        dict_ad['加购率'] = str(round(dict_ad['加购数']/dict_ad['点击量']*100)) + '%'

    
    
sales = 0
for row in ws_sl.iter_rows(min_row=2,values_only=True):
    sales += row[2]*row[3]

dict_ad['成交额'] = sales
dict_ad['成交单数'] = ws_sl.max_row-1
dict_ad['平均客单价'] = round(dict_ad['成交额']/dict_ad['成交单数'])
dict_ad['转化率%'] = str(round(dict_ad['成交单数']/dict_ad['点击量']*100,2)) + '%'
dict_ad['成交ROI'] = round(dict_ad['成交额']/dict_ad['花费'],2)

# print(dict_ad)


for col in ws_ad.iter_cols(min_row=3,min_col=3):
    # print(col)
    for i in col:
        
        if i.value == None:
            i.value = dict_ad[col[0].value]
            break

wb_ad.save('实战演练\广告\工作\{}广告投放数据.xlsx'.format(ws_ad['B'+str(i.row)].value))